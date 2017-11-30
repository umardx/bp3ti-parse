#!/usr/bin/python
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from datetime import datetime

wb		= Workbook()
now		= datetime.now()
ws		= wb.active

judul		= "Recap Sensot Uptime - " + `now.year`
operator	= "ISP Aplikanusa Lintasarta"
kontrak		= "22/PKS-AI/BP3TI/KOMINFO/01/2017"

# Format:
f_title		= Font(bold=True, size=16)
f_hcenter	= Alignment(horizontal="center")
f_vcenter	= Alignment(vertical="center") 

for val in xrange(1, 4):
	cell = "A" + `val` + ":D" + `val`
	ws.merge_cells(cell)

alphabet = map(chr, range(65,  70))
for val in alphabet:
	cell = val + "5:" + val + "6"
	ws.merge_cells(cell)
	cell = val + "5"
	ws[cell].alignment = f_vcenter

ws.merge_cells('F5:Y5')

alphabet = map(chr, range(70,  90))
numb = 1
for i, val in enumerate(alphabet):
	if i % 2 == 0:
		cell = val + "6"
		ws[cell] = numb
		numb += 1
		ws[cell].alignment = f_hcenter
		cell = cell + ":" + alphabet[i+1] + "6"
		ws.merge_cells(cell)

ws['A1'].font = f_title
ws['F5'].alignment = f_hcenter

ws['A1'] = judul
ws['A3'] = operator
ws['A4'] = kontrak

ws['A5'] = "No."
ws['B5'] = "Lokasi"
ws['C5'] = "Sensor ID"
ws['D5'] = "Start Layanan"
ws['E5'] = "Uptime Date"
ws['F5'] = "Bulan ke (Restitusi s | SLA %)"

# Save the file
wb.save("sample-out.xlsx")